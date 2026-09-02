#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Work Type 08: Excel Tax Invoice Creator
=========================================
Generates 100% vector ReportLab PDF Tax Invoices matching the exact Excel template
layout from 'Tax Invoice Formate Updates.xlsm' (Image 1 reference design).

Layout Specifications:
- Exact A4 page geometry with collinear grid lines.
- 2-column header section ('Invoice Details' vs 'Bill To') with gray header bars (#D9D9D9).
- Items table with 7 columns (Sr No, Project, HSN Code, Qty, Rate, GST %, Taxable Amount).
- Bottom summary panel with Amount in Words, Tax Breakdown Table, Bank Details, and Signature block.
"""

import os
import sys
import argparse
import openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.lib.colors import HexColor
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
try:
    from num2words import num2words
except ImportError:
    num2words = None

PAGE_W, PAGE_H = A4

# Register standard fonts
try:
    pdfmetrics.registerFont(TTFont('Arial',           'C:/Windows/Fonts/arial.ttf'))
    pdfmetrics.registerFont(TTFont('Arial-Bold',      'C:/Windows/Fonts/arialbd.ttf'))
    pdfmetrics.registerFont(TTFont('Arial-Italic',    'C:/Windows/Fonts/ariali.ttf'))
    pdfmetrics.registerFont(TTFont('Arial-BoldItalic','C:/Windows/Fonts/arialbi.ttf'))
    R, B, RI, BI = 'Arial', 'Arial-Bold', 'Arial-Italic', 'Arial-BoldItalic'
except Exception:
    R, B, RI, BI = 'Helvetica', 'Helvetica-Bold', 'Helvetica-Oblique', 'Helvetica-BoldOblique'


def format_currency(val):
    """Formats float/int into Indian numbering comma style (e.g. 1,00,000.00 or 10,77,19,741.36)."""
    try:
        fval = float(val)
        is_negative = fval < 0
        fval = abs(fval)
        
        s = f"{fval:.2f}"
        int_part, dec_part = s.split(".")
        
        if len(int_part) <= 3:
            res_int = int_part
        else:
            last3 = int_part[-3:]
            rest = int_part[:-3]
            groups = []
            while len(rest) > 2:
                groups.insert(0, rest[-2:])
                rest = rest[:-2]
            if rest:
                groups.insert(0, rest)
            res_int = ",".join(groups) + "," + last3
            
        formatted = f"{res_int}.{dec_part}"
        return f"-{formatted}" if is_negative else formatted
    except (ValueError, TypeError):
        return str(val or "")


def clean_str(val):
    """Sanitizes text string by removing unprintable/newline characters that cause square boxes in ReportLab."""
    if val is None:
        return ""
    s = str(val).replace('\n', ' ').replace('\r', ' ').replace('\t', ' ').replace('\xa0', ' ')
    return " ".join(s.split())


def clean_words_string(words):
    """Removes all commas and hyphens/dashes from Amount in Words text."""
    if not words:
        return ""
    w = clean_str(words).replace(",", "").replace("-", " ")
    return " ".join(w.split())


def draw_wrapped_text(c, text, x, y, max_width, font_name, font_size, leading=11, align="left"):
    """Draws multi-line wrapped text strictly within specified max_width budget, ensuring no text ever overflows column lines."""
    import re
    c.setFont(font_name, font_size)
    raw = clean_str(text)
    if not raw:
        return 0

    # 1. Pre-insert spaces after slashes, commas, hyphens if missing, and split CamelCase
    formatted_text = re.sub(r'([/,])(?!\s)', r'\1 ', raw)
    formatted_text = re.sub(r'([a-z])([A-Z])', r'\1 \2', formatted_text)

    words = formatted_text.split()

    # 2. Break down any individual word that is wider than max_width into sub-word chunks
    processed_words = []
    for w in words:
        if c.stringWidth(w, font_name, font_size) <= max_width:
            processed_words.append(w)
        else:
            # Word exceeds column width — split character by character
            chunk = ""
            for char in w:
                if c.stringWidth(chunk + char, font_name, font_size) <= max_width:
                    chunk += char
                else:
                    if chunk:
                        processed_words.append(chunk)
                    chunk = char
            if chunk:
                processed_words.append(chunk)

    lines = []
    curr_line = []
    for w in processed_words:
        test_line = " ".join(curr_line + [w])
        if c.stringWidth(test_line, font_name, font_size) <= max_width:
            curr_line.append(w)
        else:
            if curr_line:
                lines.append(" ".join(curr_line))
            curr_line = [w]
    if curr_line:
        lines.append(" ".join(curr_line))

    curr_y = y
    for line in lines:
        if align == "center":
            c.drawCentredString(x + max_width / 2.0, curr_y, line)
        elif align == "right":
            c.drawRightString(x + max_width, curr_y, line)
        else:
            c.drawString(x, curr_y, line)
        curr_y -= leading

    return len(lines) * leading


def draw_excel_tax_invoice(output_pdf: str, invoice_data: dict):
    """Renders clean ReportLab PDF Tax Invoice matching Reference Excel layout exactly."""
    c = canvas.Canvas(output_pdf, pagesize=A4)
    
    # ── Page Geometry & Outer Borders ─────────────────────────────────────────
    # Margins: 42pt = 14.8mm safe print zone on all 4 sides (avoids printer clip)
    X_L, X_R = 42.0, PAGE_W - 42.0   # 42 → 553  (width = 511pt)
    Y_B, Y_T = 42.0, PAGE_H - 42.0   # 42 → 800  (height = 758pt)
    BOX_W = X_R - X_L                 # 511.0 pt usable width
    
    # ── Fix 9: Use thin 0.5pt lines for all section dividers ─────────────────
    DIVIDER_W = 0.5
    BORDER_W = 1.0

    # ── Document Title ("Tax Invoice" or "Proforma Invoice") — underlined, 14pt ──
    c.setFont(B, 14)
    doc_type = str(invoice_data.get("doc_type", "")).lower().strip()
    if doc_type in ["proforma", "proforma_invoice", "proforma-invoice"]:
        title_text = "Proforma Invoice"
    else:
        title_text = str(invoice_data.get("doc_title", "Tax Invoice"))
    
    title_x = PAGE_W / 2.0
    title_y = Y_T + 10.0
    title_w = c.stringWidth(title_text, B, 14)
    c.drawCentredString(title_x, title_y, title_text)
    # Draw dynamic underline
    c.setLineWidth(0.8)
    c.line(title_x - title_w / 2.0, title_y - 2.5, title_x + title_w / 2.0, title_y - 2.5)
    
    # Outer Border Box
    c.setLineWidth(BORDER_W)
    c.setStrokeColor(HexColor("#000000"))
    c.rect(X_L, Y_B, BOX_W, Y_T - Y_B)

    # ── 1. Supplier Company Header ───────────────────────────────────────────
    # Fix 2: Reduce address font to 7.5pt, add 18pt padding before Invoice Details
    Y_HDR_BOTTOM = Y_T - 89.0    # proportional: was 814-89=725, now 800-89=711
    
    c.setFont(B, 14)
    supplier_name = clean_str(invoice_data.get("supplier_name", "Shivam Builders"))
    c.drawCentredString(PAGE_W / 2.0, Y_T - 22.0, supplier_name)
    
    c.setFont(R, 7.5)  # Fix 2: reduced from 8.5 → 7.5
    supplier_addr = clean_str(invoice_data.get("supplier_addr", "290, THE MEADOWS, GOKULDHAM,NR EKLAVYA SCHOOL,SARKHEJ-SANAND ROAD, AHMEDABAD, GUJARAT, 382210"))
    c.drawCentredString(PAGE_W / 2.0, Y_T - 36.0, supplier_addr)
    
    c.setFont(B, 9)
    supplier_gst = clean_str(invoice_data.get("supplier_gst", "24ABDFS4611H1ZG"))
    c.drawCentredString(PAGE_W / 2.0, Y_T - 49.0, f"GSTIN NO : -  {supplier_gst}")
    
    c.setLineWidth(DIVIDER_W)  # Fix 9: thin divider
    c.line(X_L, Y_HDR_BOTTOM, X_R, Y_HDR_BOTTOM)

    # ── 2. Invoice Details vs Bill To Box (2 Columns) ─────────────────────────
    Y_BILLTO_BOTTOM = Y_T - 169.0     # proportional: was 814-169=645, now 800-169=631
    X_MID = X_L + 228.0               # proportional center split (was 270-28=242 offset, scaled to 511/539)
    
    # Vertical line dividing left and right columns
    c.setLineWidth(DIVIDER_W)
    c.line(X_MID, Y_HDR_BOTTOM, X_MID, Y_BILLTO_BOTTOM)
    c.line(X_L, Y_BILLTO_BOTTOM, X_R, Y_BILLTO_BOTTOM)
    
    # Gray Header Bars
    c.setFillColor(HexColor("#D9D9D9"))
    c.rect(X_L, Y_HDR_BOTTOM - 14.0, X_MID - X_L, 14.0, fill=1, stroke=1)
    c.rect(X_MID, Y_HDR_BOTTOM - 14.0, X_R - X_MID, 14.0, fill=1, stroke=1)
    
    c.setFillColor(HexColor("#000000"))
    c.setFont(B, 8.5)
    c.drawString(X_L + 5.0, Y_HDR_BOTTOM - 11.0, "Invoice Details : -")
    c.drawString(X_MID + 5.0, Y_HDR_BOTTOM - 11.0, "Bill To : -")

    # Left Column: Invoice Details
    y_detail = Y_HDR_BOTTOM - 26.0
    c.setFont(B, 8.5)
    c.drawString(X_L + 5.0, y_detail, "Invoice No: ")
    c.setFont(R, 8.5)
    c.drawString(X_L + 62.0, y_detail, str(invoice_data.get("inv_no", "RA BILL 1")))

    y_detail -= 12.0
    c.setFont(B, 8.5)
    c.drawString(X_L + 5.0, y_detail, "Invoice Date: ")
    c.setFont(R, 8.5)
    c.drawString(X_L + 67.0, y_detail, str(invoice_data.get("inv_date", "07/07/2026")))

    y_detail -= 12.0
    c.setFont(B, 8.5)
    c.drawString(X_L + 5.0, y_detail, "State: ")
    c.setFont(R, 8.5)
    c.drawString(X_L + 34.0, y_detail, str(invoice_data.get("state", "Gujarat")))

    y_detail -= 12.0
    c.setFont(B, 8.5)
    c.drawString(X_L + 5.0, y_detail, "State Code: ")
    c.setFont(R, 8.5)
    c.drawString(X_L + 57.0, y_detail, str(invoice_data.get("state_code", "24")))

    # Right Column: Bill To
    c.setFont(B, 9)
    cust_name = str(invoice_data.get("cust_name", "Ahmedabad Municipal Corporation"))
    c.drawString(X_MID + 5.0, Y_HDR_BOTTOM - 26.0, cust_name)

    cust_addr = str(invoice_data.get("cust_addr", "AHMEDABAD MUNICIPAL CORPORATION, MAHANAGAR SEVA SADAN, SARDAR PATEL BHAVAN, DANAPITH, AHMEDABAD - 380001."))
    draw_wrapped_text(c, cust_addr, X_MID + 5.0, Y_HDR_BOTTOM - 38.0, (X_R - X_MID - 10.0), R, 7.5, leading=9.5)

    c.setFont(B, 8.5)
    c.drawString(X_MID + 5.0, Y_BILLTO_BOTTOM + 5.0, str(invoice_data.get("cust_gst", "24AAALA0024C3Z7")))

    # ── 3. Items Table Headers ───────────────────────────────────────────────
    # Columns proportionally rescaled from 539pt → 511pt usable width
    # Scale factor: 511/539 ≈ 0.948. Offsets from X_L=42:
    # Old: [28,70,270,320,355,430,475,567] → subtract 28, scale ×511/539, add 42
    COL_X = [
        X_L,                    # 42  (was 28)
        X_L + 40.0,             # 82  (was 70)
        X_L + 226.0,            # 268 (was 270, near same)
        X_L + 274.0,            # 316 (was 320)
        X_L + 308.0,            # 350 (was 355)
        X_L + 379.0,            # 421 (was 430)
        X_L + 422.0,            # 464 (was 475)
        X_R,                    # 553 (was 567)
    ]
    Y_TABLE_HEADER_TOP = Y_BILLTO_BOTTOM
    Y_TABLE_HEADER_BOTTOM = Y_BILLTO_BOTTOM - 18.0
    
    # Gray header background
    c.setFillColor(HexColor("#D9D9D9"))
    c.rect(X_L, Y_TABLE_HEADER_BOTTOM, BOX_W, 18.0, fill=1, stroke=1)
    
    c.setFillColor(HexColor("#000000"))
    c.setFont(B, 8.5)
    c.drawCentredString((COL_X[0] + COL_X[1]) / 2.0, Y_TABLE_HEADER_BOTTOM + 5.0, "Sr No.")
    c.drawCentredString((COL_X[1] + COL_X[2]) / 2.0, Y_TABLE_HEADER_BOTTOM + 5.0, "Project")
    c.drawCentredString((COL_X[2] + COL_X[3]) / 2.0, Y_TABLE_HEADER_BOTTOM + 5.0, "HSN Code")
    c.drawCentredString((COL_X[3] + COL_X[4]) / 2.0, Y_TABLE_HEADER_BOTTOM + 5.0, "Qty")
    c.drawCentredString((COL_X[4] + COL_X[5]) / 2.0, Y_TABLE_HEADER_BOTTOM + 5.0, "Rate")
    c.drawCentredString((COL_X[5] + COL_X[6]) / 2.0, Y_TABLE_HEADER_BOTTOM + 5.0, "GST %")
    c.drawCentredString((COL_X[6] + COL_X[7]) / 2.0, Y_TABLE_HEADER_BOTTOM + 5.0, "Taxable Amount")

    # ── 4. Items Table Data & Vertical Grid Lines ─────────────────────────────
    # Bottom of items area: leave proportional space for totals + bank details
    Y_WORDS_BAR_TOP = Y_B + 278.0   # proportional: was 320-28=292 above bottom, keep 278 above Y_B=42
    
    # Draw vertical grid lines for table columns
    c.setLineWidth(DIVIDER_W)
    for x in COL_X[1:-1]:
        c.line(x, Y_TABLE_HEADER_TOP, x, Y_WORDS_BAR_TOP)

    # Populate Line Items
    items = invoice_data.get("items", [{
        "sr": 1,
        "project": "Tender Documents For Laying of Water Distribution Network, Providing & Laying of Sewerage Network with Demolition & Construction of Road for Laying Network in Asarwa Ward (Part-IV) in Central Zone of AMC area",
        "hsn": "9954",
        "qty": "1.00",
        "rate": invoice_data.get("taxable_amount", 44110882.20),
        "gst_pct": "18%",
        "taxable": invoice_data.get("taxable_amount", 44110882.20)
    }])

    curr_item_y = Y_TABLE_HEADER_BOTTOM - 18.0
    for item in items:
        first_line_y = curr_item_y  # Fix 4: baseline alignment anchor

        # Fix 3: Project description LEFT-aligned, not center
        c.setFont(R, 7.5)
        proj_text = str(item.get("project", ""))
        draw_wrapped_text(c, proj_text, COL_X[1] + 6.0, first_line_y,
                          (COL_X[2] - COL_X[1] - 12.0), R, 7.5, leading=10, align="left")

        # Fix 4: All number columns share same Y baseline as first project text line
        c.setFont(R, 8.5)
        c.drawCentredString((COL_X[0] + COL_X[1]) / 2.0, first_line_y, str(item.get("sr", 1)))
        c.drawCentredString((COL_X[2] + COL_X[3]) / 2.0, first_line_y, str(item.get("hsn", "9954")))
        c.drawCentredString((COL_X[3] + COL_X[4]) / 2.0, first_line_y, str(item.get("qty", "1")))
        c.drawRightString(COL_X[5] - 6.0, first_line_y, format_currency(item.get("rate", 0)))
        c.drawCentredString((COL_X[5] + COL_X[6]) / 2.0, first_line_y, str(item.get("gst_pct", "18%")))
        c.drawRightString(COL_X[7] - 6.0, first_line_y, format_currency(item.get("taxable", 0)))

    # ── 5. Sub Total Bar & Amount in Words Header Bar ────────────────────────
    X_SUMMARY_SPLIT = COL_X[2]  # 270.0 pt (Collinear with Project / HSN Code column line as per reference layout)
    Y_SUBTOTAL_BAR_H = 15.0
    
    c.setLineWidth(DIVIDER_W)
    c.line(X_L, Y_WORDS_BAR_TOP, X_R, Y_WORDS_BAR_TOP)
    
    # Gray background on BOTH Left side ("Total Invoice Amount in words : -") AND Right side ("Sub Total : -")
    c.setFillColor(HexColor("#D9D9D9"))
    c.rect(X_L, Y_WORDS_BAR_TOP - Y_SUBTOTAL_BAR_H, (X_SUMMARY_SPLIT - X_L), Y_SUBTOTAL_BAR_H, fill=1, stroke=1)
    c.rect(X_SUMMARY_SPLIT, Y_WORDS_BAR_TOP - Y_SUBTOTAL_BAR_H, (X_R - X_SUMMARY_SPLIT), Y_SUBTOTAL_BAR_H, fill=1, stroke=1)
    
    c.setFillColor(HexColor("#000000"))
    c.setFont(B, 8.5)
    c.drawCentredString((X_L + X_SUMMARY_SPLIT) / 2.0, Y_WORDS_BAR_TOP - 11.0, "Total Invoice Amount in words : -")
    c.drawString(X_SUMMARY_SPLIT + 10.0, Y_WORDS_BAR_TOP - 11.0, "Sub Total : -")
    
    taxable_val = float(invoice_data.get("taxable_amount", 44110882.20))
    c.drawRightString(X_R - 6.0, Y_WORDS_BAR_TOP - 11.0, format_currency(taxable_val))

    # ── 6. Tax Breakdown Rows (Right Panel) — WHITE background ───────────────
    Y_SUBTOTAL_BOTTOM = Y_WORDS_BAR_TOP - Y_SUBTOTAL_BAR_H
    
    cgst_val = float(invoice_data.get("cgst", taxable_val * 0.09))
    sgst_val = float(invoice_data.get("sgst", taxable_val * 0.09))
    subtotal_before_round = taxable_val + cgst_val + sgst_val
    
    # Determine Round Off & Grand Total (Whole Number)
    if "round_off" in invoice_data and invoice_data["round_off"] is not None:
        round_off_val = float(invoice_data["round_off"])
        grand_val = float(invoice_data.get("grand_total", subtotal_before_round + round_off_val))
    elif "grand_total" in invoice_data and invoice_data["grand_total"] is not None:
        grand_val = float(invoice_data["grand_total"])
        round_off_val = round(grand_val - subtotal_before_round, 2)
    else:
        # Default: Auto Round-Off to nearest whole integer (0 paisa)
        nearest_whole = round(subtotal_before_round)
        round_off_val = round(nearest_whole - subtotal_before_round, 2)
        grand_val = float(nearest_whole)

    summary_rows = [
        ("Taxable  Amount : -", format_currency(taxable_val), False),
        ("CGST 9% : -", format_currency(cgst_val), False),
        ("SGST 9% : -", format_currency(sgst_val), False),
        ("Total : -", format_currency(subtotal_before_round), True),
    ]

    if abs(round_off_val) >= 0.01:
        summary_rows.append(("Round Off : -", f"{round_off_val:+.2f}", False))

    summary_rows.append(("Grand Total : -", format_currency(grand_val), True))

    # Reduce row height from 24pt → 20pt
    row_h = 20.0
    X_VAL_SPLIT = COL_X[6]  # 475.0 pt (Collinear with GST % / Taxable Amount column line)
    curr_sum_y = Y_SUBTOTAL_BOTTOM
    
    # Vertical divider for the tax breakdown column
    Y_SUMMARY_BOTTOM_CALC = Y_SUBTOTAL_BOTTOM - (len(summary_rows) * row_h)
    c.setLineWidth(DIVIDER_W)
    c.line(X_SUMMARY_SPLIT, Y_WORDS_BAR_TOP, X_SUMMARY_SPLIT, Y_SUMMARY_BOTTOM_CALC)
    c.line(X_VAL_SPLIT, Y_SUBTOTAL_BOTTOM, X_VAL_SPLIT, Y_SUMMARY_BOTTOM_CALC)

    for label, val_str, is_bold in summary_rows:
        curr_sum_y -= row_h

        # Only Grand Total row gets gray, rest is white
        if label.startswith("Grand Total"):
            c.setFillColor(HexColor("#D9D9D9"))
            c.rect(X_SUMMARY_SPLIT, curr_sum_y, (X_R - X_SUMMARY_SPLIT), row_h, fill=1, stroke=1)
            c.setFillColor(HexColor("#000000"))

        f_font = B if is_bold else R
        c.setFont(f_font, 8.5)
        c.drawRightString(X_VAL_SPLIT - 6.0, curr_sum_y + 6.0, label)
        c.drawRightString(X_R - 6.0, curr_sum_y + 6.0, val_str)

        c.setLineWidth(DIVIDER_W)
        c.line(X_SUMMARY_SPLIT, curr_sum_y, X_R, curr_sum_y)

    Y_SUMMARY_BOTTOM = curr_sum_y

    # ── 7. Amount in Words (Left Panel) with Box Border & Clean Framing ────────
    words_raw = str(invoice_data.get("words", "Five Crore Twenty Lakh Fifty Thousand Eight Hundred Forty One Rupees Only"))
    words_str = clean_words_string(words_raw)
    words_box_top = Y_SUBTOTAL_BOTTOM
    words_box_bottom = Y_SUMMARY_BOTTOM
    words_box_height = words_box_top - words_box_bottom
    
    # Draw explicit box border for Amount in Words box
    c.setLineWidth(DIVIDER_W)
    c.rect(X_L, words_box_bottom, (X_SUMMARY_SPLIT - X_L), words_box_height, fill=0, stroke=1)
    
    # Estimate number of wrapped lines
    c.setFont(B, 8.5)
    words_list = str(words_str).split()
    est_lines = []
    curr_wline = []
    avail_w = X_SUMMARY_SPLIT - X_L - 30.0
    for w in words_list:
        test = " ".join(curr_wline + [w])
        if c.stringWidth(test, B, 8.5) <= avail_w:
            curr_wline.append(w)
        else:
            if curr_wline:
                est_lines.append(" ".join(curr_wline))
            curr_wline = [w]
    if curr_wline:
        est_lines.append(" ".join(curr_wline))
    
    text_block_h = len(est_lines) * 12.0
    words_y_start = words_box_bottom + (words_box_height - text_block_h) / 2.0 + text_block_h - 12.0
    
    c.setFillColor(HexColor("#000000"))
    draw_wrapped_text(c, words_str, X_L + 15.0, words_y_start,
                      avail_w, B, 8.5, leading=12, align="center")

    # Bottom boundary of summary section
    c.setLineWidth(DIVIDER_W)
    c.line(X_L, Y_SUMMARY_BOTTOM, X_R, Y_SUMMARY_BOTTOM)

    # ── 8. Bank Details & Signature Section ──────────────────────────────────
    X_BANK_SPLIT = COL_X[2]  # 270.0 pt (Collinear with Project / HSN Code column line as per reference layout)
    c.setLineWidth(DIVIDER_W)
    c.line(X_BANK_SPLIT, Y_SUMMARY_BOTTOM, X_BANK_SPLIT, Y_B)
    
    # Left Box Header: Bank Details
    c.setFillColor(HexColor("#D9D9D9"))
    c.rect(X_L, Y_SUMMARY_BOTTOM - 14.0, (X_BANK_SPLIT - X_L), 14.0, fill=1, stroke=1)
    c.setFillColor(HexColor("#000000"))
    c.setFont(B, 8.5)
    c.drawString(X_L + 5.0, Y_SUMMARY_BOTTOM - 11.0, "Bank Details : -")

    # Bank Metadata fields
    y_bank = Y_SUMMARY_BOTTOM - 26.0
    c.setFont(B, 8.0)
    c.drawString(X_L + 5.0, y_bank, "Account No : - ")
    c.setFont(R, 8.0)
    c.drawString(X_L + 75.0, y_bank, str(invoice_data.get("acc_no", "42350500563")))

    y_bank -= 12.0
    c.setFont(B, 8.0)
    c.drawString(X_L + 5.0, y_bank, "Bank Name : - ")
    c.setFont(R, 8.0)
    c.drawString(X_L + 75.0, y_bank, str(invoice_data.get("bank_name", "ICICI BANK")))

    y_bank -= 12.0
    c.setFont(B, 8.0)
    c.drawString(X_L + 5.0, y_bank, "Branch : - ")
    c.setFont(R, 8.0)
    c.drawString(X_L + 75.0, y_bank, str(invoice_data.get("branch", "MAKARBA")))

    y_bank -= 12.0
    c.setFont(B, 8.0)
    c.drawString(X_L + 5.0, y_bank, "IFSC Code : - ")
    c.setFont(R, 8.0)
    c.drawString(X_L + 75.0, y_bank, str(invoice_data.get("ifsc", "ICIC0004235")))

    # Terms & Conditions Gray Header Bar
    y_terms = y_bank - 8.0
    c.setFillColor(HexColor("#D9D9D9"))
    c.rect(X_L, y_terms - 14.0, (X_BANK_SPLIT - X_L), 14.0, fill=1, stroke=1)
    c.setFillColor(HexColor("#000000"))
    c.setFont(B, 8.5)
    c.drawString(X_L + 5.0, y_terms - 11.0, "Terms & Conditions : -")

    # Terms & Conditions List Points
    terms_list = invoice_data.get("terms", [
        "1. Goods/Services once billed will not be taken back.",
        "2. Subject to Ahmedabad Jurisdiction only.",
        "3. E. & O.E."
    ])
    curr_t_y = y_terms - 25.0
    c.setFont(R, 7.0)
    for term in terms_list:
        c.drawString(X_L + 5.0, curr_t_y, str(term))
        curr_t_y -= 10.0

    # Right Box: Signature Block — Authentic Physical Hand-Stamp Effect
    sig_center_x = (X_BANK_SPLIT + X_R) / 2.0  # 433.5 pt
    y_top_text = Y_SUMMARY_BOTTOM - 26.0
    y_bottom_text = Y_B + 14.0
    y_mid_section = (y_top_text + y_bottom_text) / 2.0

    # Draw signature block text first
    c.setFillColor(HexColor("#000000"))
    c.setFont(B, 9)
    c.drawCentredString(sig_center_x, y_top_text, f"For, {supplier_name}")
    c.setFont(B, 9)
    c.drawCentredString(sig_center_x, y_bottom_text, "Authorised Signatory")
    
    # Auto-detect or load stamp image from Stamps directory if include_stamp is True (Default: True)
    include_stamp = invoice_data.get("include_stamp", True)
    stamp_path = invoice_data.get("stamp_path")
    
    possible_stamps_dirs = [
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "Stamps"),
        os.path.join(os.getcwd(), "Stamps"),
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "Stamps"),
    ]
    stamps_dir = next((sd for sd in possible_stamps_dirs if os.path.exists(sd)), None)
    
    if include_stamp and not stamp_path and stamps_dir and os.path.exists(stamps_dir):
        safe_supplier = "".join(c for c in supplier_name if c.isalnum() or c in (" ", "_", "-")).strip().replace(" ", "_")
        exact_png = os.path.join(stamps_dir, f"{safe_supplier}.png")
        exact_jpg = os.path.join(stamps_dir, f"{safe_supplier}.jpg")
        
        if os.path.exists(exact_png):
            stamp_path = exact_png
        elif os.path.exists(exact_jpg):
            stamp_path = exact_jpg
        else:
            sup_lower = supplier_name.lower()
            for fname in os.listdir(stamps_dir):
                f_lower = fname.lower()
                if not f_lower.endswith((".png", ".jpg", ".jpeg")):
                    continue
                if "shivam" in sup_lower and "shivam" in f_lower:
                    stamp_path = os.path.join(stamps_dir, fname); break
                elif "yogi" in sup_lower and "yogi" in f_lower:
                    stamp_path = os.path.join(stamps_dir, fname); break
                elif "khodiyar" in sup_lower and "khodiyar" in f_lower:
                    stamp_path = os.path.join(stamps_dir, fname); break
                elif "varudi" in sup_lower and "varudi" in f_lower:
                    stamp_path = os.path.join(stamps_dir, fname); break
                elif "sarthi" in sup_lower and "sarthi" in f_lower:
                    stamp_path = os.path.join(stamps_dir, fname); break
                elif "jnp" in sup_lower and "jnp" in f_lower:
                    stamp_path = os.path.join(stamps_dir, fname); break

    # Draw stamp image OVER the text with realistic hand-stamped tilt (-3 deg) and natural overlap
    if include_stamp and stamp_path and os.path.exists(stamp_path):
        stamp_w, stamp_h = 100.0, 100.0
        stamp_cy = y_mid_section - 2.0  # Slightly lower for natural overlap over Authorised Signatory
        try:
            c.saveState()
            c.translate(sig_center_x, stamp_cy)
            c.rotate(-3.5)  # Slight 3.5° physical hand-stamp tilt
            c.drawImage(stamp_path, -stamp_w / 2.0, -stamp_h / 2.0, width=stamp_w, height=stamp_h, mask='auto', preserveAspectRatio=True)
            c.restoreState()
        except Exception as e:
            c.restoreState()
            print(f"  [WARNING] Stamp image draw failed: {e}")

    c.save()


def parse_excel_and_generate(excel_path: str, output_dir: str):
    """Parses workbook and renders PDFs matching Image 1 layout."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    os.makedirs(output_dir, exist_ok=True)
    
    sheet_names = [s for s in ['Total Amount', 'Basic Amount'] if s in wb.sheetnames]
    if not sheet_names:
        sheet_names = wb.sheetnames[:1]

    for sheet_name in sheet_names:
        s = wb[sheet_name]
        
        # Read exact cell values from Image 1 Excel layout
        supplier_name = s.cell(row=8, column=4).value or "Shivam Builders"
        supplier_addr = s.cell(row=9, column=4).value or "290, THE MEADOWS, GOKULDHAM,NR EKLAVYA SCHOOL,SARKHEJ-SANAND ROAD, AHMEDABAD, GUJARAT, 382210"
        supplier_gst = str(s.cell(row=10, column=4).value or "").replace("GSTIN NO : -", "").strip() or "24ABDFS4611H1ZG"

        inv_no_str = str(s.cell(row=15, column=4).value or "Invoice No: RA BILL 1").replace("Invoice No:", "").strip()
        inv_date_str = str(s.cell(row=16, column=4).value or "Invoice Date: 07/07/2026").replace("Invoice Date:", "").strip()

        cust_name = s.cell(row=15, column=6).value or "Ahmedabad Municipal Corporation"
        cust_addr = s.cell(row=16, column=6).value or "AHMEDABAD MUNICIPAL CORPORATION, MAHANAGAR SEVA SADAN, SARDAR PATEL BHAVAN, DANAPITH, AHMEDABAD - 380001."
        cust_gst = str(s.cell(row=18, column=6).value or "24AAALA0024C3Z7")

        proj_desc = s.cell(row=20, column=5).value or "Tender Documents For Laying of Water Distribution Network..."
        hsn_code = str(s.cell(row=20, column=6).value or "9954")
        qty_val = str(s.cell(row=20, column=7).value or "1.00")
        rate_val = s.cell(row=20, column=8).value or 44110882.20
        gst_pct_val = "18%"
        taxable_val = s.cell(row=20, column=10).value or rate_val

        words_val = s.cell(row=34, column=4).value or "Five Crore Twenty Lakh Fifty Thousand Eight Hundred Forty One Rupees Only"

        cgst_val = s.cell(row=35, column=10).value or (taxable_val * 0.09)
        sgst_val = s.cell(row=36, column=10).value or (taxable_val * 0.09)
        total_val = s.cell(row=37, column=10).value or (taxable_val + cgst_val + sgst_val)
        grand_val = s.cell(row=39, column=10).value or total_val

        acc_no = str(s.cell(row=40, column=5).value or "42350500563")
        bank_name = str(s.cell(row=41, column=5).value or "ICICI BANK")
        branch_name = str(s.cell(row=42, column=5).value or "MAKARBA")
        ifsc_code = str(s.cell(row=43, column=5).value or "ICIC0004235")

        data = {
            "supplier_name": supplier_name,
            "supplier_addr": supplier_addr,
            "supplier_gst": supplier_gst,
            "inv_no": inv_no_str,
            "inv_date": inv_date_str,
            "state": "Gujarat",
            "state_code": "24",
            "cust_name": cust_name,
            "cust_addr": cust_addr,
            "cust_gst": cust_gst,
            "items": [{
                "sr": 1,
                "project": proj_desc,
                "hsn": hsn_code,
                "qty": qty_val,
                "rate": rate_val,
                "gst_pct": gst_pct_val,
                "taxable": taxable_val
            }],
            "words": words_val,
            "taxable_amount": taxable_val,
            "cgst": cgst_val,
            "sgst": sgst_val,
            "total": total_val,
            "grand_total": grand_val,
            "acc_no": acc_no,
            "bank_name": bank_name,
            "branch": branch_name,
            "ifsc": ifsc_code
        }

        sanitized_name = sheet_name.lower().replace(" ", "_")
        out_pdf = os.path.join(output_dir, f"tax_invoice_{sanitized_name}.pdf")
        draw_excel_tax_invoice(out_pdf, data)
        print(f"  [SUCCESS] Generated Image 1 Matched PDF: {out_pdf}")


def main():
    parser = argparse.ArgumentParser(description="Excel Tax Invoice Creator (Image 1 Layout)")
    parser.add_argument("--excel", default="08_Excel_Tax_Invoice_Creator/Tax Invoice Formate Updates.xlsm", help="Path to Excel workbook")
    parser.add_argument("--output-dir", default="08_Excel_Tax_Invoice_Creator/outputs", help="Output directory for generated PDFs")
    args = parser.parse_args()

    parse_excel_and_generate(args.excel, args.output_dir)


if __name__ == "__main__":
    main()
